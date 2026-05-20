"""
seed.py — One-time seeding of reference tables from spreadsheets.

Requirements:
    pip install openpyxl supabase python-dotenv

Usage:
    python scripts/seed.py

Run from the project root (fcproject2026/).
The script reads .env.local for SUPABASE credentials.
"""

import os
import sys
from pathlib import Path
from dotenv import load_dotenv
import openpyxl
from supabase import create_client, Client

# ── Config ──────────────────────────────────────────────────────────────────

load_dotenv(Path(__file__).parent.parent / ".env.local")

SUPABASE_URL = os.environ["NEXT_PUBLIC_SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]

SPREADSHEET_DIR = Path(r"C:\Users\kimbe\OneDrive - Quikway Group\FC Project")

def get_client() -> Client:
    return create_client(SUPABASE_URL, SUPABASE_KEY)


def rows(sheet, header_row: int = 1):
    """Yield each data row as a dict keyed by header name."""
    headers = [cell.value for cell in sheet[header_row]]
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if all(v is None for v in row):
            continue
        yield dict(zip(headers, row))


def clean_str(v) -> str | None:
    if v is None:
        return None
    return str(v).strip() or None


def to_bool(v) -> bool:
    return bool(v) if v is not None else False


# ── Seed: fuel_city_site_details → sites ────────────────────────────────────

def seed_sites(sb: Client):
    wb = openpyxl.load_workbook(SPREADSHEET_DIR / "fuel_city_site_details.xlsx", data_only=True)
    ws = wb.active
    records = []
    for r in rows(ws):
        records.append({
            "site_id":          int(r["Site_ID"]),
            "site_name":        clean_str(r["Site_Name"]),
            "site_address":     clean_str(r["Site_Address"]),
            "city":             clean_str(r["City"]),
            "state":            clean_str(r["State"]),
            "zip":              clean_str(r.get("zip") or r.get("Zip")),
            "longitude":        float(r["Longitude"]) if r["Longitude"] else None,
            "latitude":         float(r["Latitude"]) if r["Latitude"] else None,
            "auto_diesel":      to_bool(r.get("Auto_diesel")),
            "truck_diesel":     to_bool(r.get("Truck_diesel")),
            "bio_tank":         to_bool(r.get("Bio_tank")),
            "auto_diesel_note": clean_str(r.get("Auto_diesel_note")),
            "truck_diesel_note":clean_str(r.get("Truck_diesel_note")),
            "bio_tank_note":    clean_str(r.get("bio_tank_note")),
        })
    res = sb.table("sites").upsert(records).execute()
    print(f"  sites: {len(records)} rows upserted")


# ── Seed: other_site_details → all_sites ────────────────────────────────────

def seed_all_sites(sb: Client):
    wb = openpyxl.load_workbook(SPREADSHEET_DIR / "other_site_details.xlsx", data_only=True)
    ws = wb.active
    records = []
    for r in rows(ws):
        if not r.get("site_id"):
            continue
        records.append({
            "site_id":             int(r["site_id"]),
            "site_name":           clean_str(r.get("site_name")),
            "site_address":        clean_str(r.get("site_address")),
            "city":                clean_str(r.get("city")),
            "state":               clean_str(r.get("state")),
            "zip":                 clean_str(r.get("zip")),
            "longitude":           float(r["longitude"]) if r.get("longitude") else None,
            "latitude":            float(r["latitude"]) if r.get("latitude") else None,
            "customer_group_name": clean_str(r.get("customer_group_name")),
        })
    res = sb.table("all_sites").upsert(records).execute()
    print(f"  all_sites: {len(records)} rows upserted")


# ── Seed: Terminal_Locations → terminals ─────────────────────────────────────

def seed_terminals(sb: Client):
    wb = openpyxl.load_workbook(SPREADSHEET_DIR / "Terminal_Locations.xlsx", data_only=True)
    ws = wb.active
    records = []
    for r in rows(ws):
        tid = clean_str(r.get("terminal_id"))
        if not tid:
            continue
        records.append({
            "terminal_id":           tid,
            "terminal_abbreviation": clean_str(r.get("terminal_abreviation")),
            "terminal_name":         clean_str(r.get("terminal_name")),
            "terminal_address":      clean_str(r.get("terminal_address")),
            "city":                  clean_str(r.get("city")),
            "state":                 clean_str(r.get("state")),
            "latitude":              float(r["latitude"]) if r.get("latitude") else None,
            "longitude":             float(r["longitude"]) if r.get("longitude") else None,
            "is_fuel_city":          to_bool(r.get("is_fuel_city")),
            "is_custom":             False,
        })
    # Deduplicate by terminal_id (spreadsheet may have duplicate rows)
    seen = {}
    for r in records:
        seen[r["terminal_id"]] = r
    records = list(seen.values())

    res = sb.table("terminals").upsert(records).execute()
    print(f"  terminals: {len(records)} rows upserted")


# ── Seed: suppliers → suppliers ─────────────────────────────────────────────

def seed_suppliers(sb: Client):
    wb = openpyxl.load_workbook(SPREADSHEET_DIR / "suppliers.xlsx", data_only=True)
    ws = wb.active
    records = []
    for r in rows(ws):
        if not r.get("Supplier_Name"):
            continue
        records.append({
            "supplier_id":             int(r["Supplier_ID"]),
            "supplier_name":           clean_str(r["Supplier_Name"]),
            "supplier_loading_number": clean_str(r["Supplier_Loading_Number"]),
        })
    res = sb.table("suppliers").upsert(records).execute()
    print(f"  suppliers: {len(records)} rows upserted")


# ── Seed: Email_Notifications → email_notifications ─────────────────────────

def seed_email_notifications(sb: Client):
    wb = openpyxl.load_workbook(SPREADSHEET_DIR / "Email_Notifications.xlsx", data_only=True)
    ws = wb.active
    records = []
    for r in rows(ws):
        if not r.get("Email"):
            continue
        records.append({
            "email_id": int(r["Email_ID"]),
            "name":     clean_str(r["Name"]),
            "email":    clean_str(r["Email"]),
            "send":     to_bool(r.get("Send")),
            "receive":  to_bool(r.get("Receive")),
            "active":   True,
        })
    res = sb.table("email_notifications").upsert(records).execute()
    print(f"  email_notifications: {len(records)} rows upserted")


# ── Seed: Loads Feed → loads (initial load) ──────────────────────────────────

def seed_loads(sb: Client):
    """Seed an initial batch of loads. The sync script will keep this up to date."""
    from datetime import datetime, timezone

    wb = openpyxl.load_workbook(SPREADSHEET_DIR / "Loads Feed for Kim.xlsx", data_only=True)
    ws = wb["Main"]
    records = []
    now = datetime.now(timezone.utc).isoformat()

    for r in rows(ws):
        if not r.get("CE_ID"):
            continue

        def ts(v):
            if v is None:
                return None
            if hasattr(v, "isoformat"):
                # already a datetime from openpyxl
                return v.isoformat()
            return str(v)

        records.append({
            "ce_id":               int(r["CE_ID"]),
            "delivery_date":       ts(r["Delivery_Date"]),
            "customer":            clean_str(r.get("Customer")),
            "order_number":        clean_str(r.get("Order_Number")),
            "site_id":             int(r["Site_ID"]) if r.get("Site_ID") else None,
            "site_name":           clean_str(r.get("Site_Name")),
            "terminal_id":         clean_str(r.get("Terminal_ID")),
            "terminal_name":       clean_str(r.get("Terminal_Name")),
            "product_name":        clean_str(r.get("Product_Name")) or "Unknown",
            "gallons_ordered":     float(r["Gallons_Ordered"]) if r.get("Gallons_Ordered") else None,
            "site_address":        clean_str(r.get("Site_Address")),
            "city":                clean_str(r.get("City")),
            "state":               clean_str(r.get("State")),
            "first_name":          clean_str(r.get("First_Name")),
            "last_name":           clean_str(r.get("Last_Name")),
            "start_window":        ts(r.get("Start_Window")),
            "end_window":          ts(r.get("End_Window")),
            "delivery_eta":        ts(r.get("Delivery_ETA")),
            "arrived_at_rack_time":ts(r.get("Arrived_at_Rack_Time")),
            "load_status":         int(r["Load_Status"]) if r.get("Load_Status") is not None else 1,
            "customer_id":         int(r["Customer_ID"]) if r.get("Customer_ID") else None,
            "synced_at":           now,
        })

    # Upsert in batches of 500
    for i in range(0, len(records), 500):
        sb.table("loads").upsert(
            records[i:i+500],
            on_conflict="ce_id,product_name"
        ).execute()

    print(f"  loads: {len(records)} rows upserted")


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    print("Connecting to Supabase...")
    sb = get_client()

    print("Seeding reference tables...")
    seed_sites(sb)
    seed_all_sites(sb)
    seed_terminals(sb)
    seed_suppliers(sb)
    seed_email_notifications(sb)

    print("Seeding initial load data...")
    try:
        seed_loads(sb)
    except PermissionError:
        print("  loads: skipped (file is open in Excel or locked by OneDrive)")
        print("  -> Run scripts/sync.py once Excel is closed to populate loads.")

    print("\nDone! All tables seeded.")
    print("\nNext steps:")
    print("  1. Run scripts/schema.sql in the Supabase SQL Editor if you haven't already.")
    print("  2. Set up the sync script (scripts/sync.py) in Windows Task Scheduler.")


if __name__ == "__main__":
    main()
