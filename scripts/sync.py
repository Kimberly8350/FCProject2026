"""
sync.py — Reads the refreshed Loads Feed Excel file and upserts into Supabase.

Run every 15 minutes via Windows Task Scheduler.
See scripts/task_scheduler_setup.md for setup instructions.

Requirements:
    pip install openpyxl supabase python-dotenv pywin32

Usage:
    python scripts/sync.py
"""

import os
import sys
import time
import logging
from pathlib import Path
from datetime import datetime, timezone, date
from zoneinfo import ZoneInfo

CENTRAL = ZoneInfo('America/Chicago')
from dotenv import load_dotenv
import openpyxl
from supabase import create_client, Client

# ── Logging ─────────────────────────────────────────────────────────────────

LOG_FILE = Path(__file__).parent / "sync.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# ── Config ───────────────────────────────────────────────────────────────────

load_dotenv(Path(__file__).parent.parent / ".env.local")

SUPABASE_URL = os.environ["NEXT_PUBLIC_SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]

LOADS_FILE = Path(r"C:\Users\kimbe\OneDrive - Quikway Group\FC Project\Loads Feed for Kim.xlsx")


def get_client() -> Client:
    return create_client(SUPABASE_URL, SUPABASE_KEY)


def clean_str(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def to_ts(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        if v.tzinfo is None:
            v = v.replace(tzinfo=CENTRAL)  # Excel times are Central Time
        return v.isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v)


def trigger_excel_refresh():
    """
    Ask Excel to refresh the workbook's data connections before we read it.
    Requires pywin32 and Excel to be installed.
    Falls back silently if Excel isn't available.
    """
    try:
        import win32com.client
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(LOADS_FILE.resolve()))
        wb.RefreshAll()
        # Wait for background queries to finish (up to 60 seconds)
        deadline = time.time() + 60
        while time.time() < deadline:
            try:
                wb.Save()
                break
            except Exception:
                time.sleep(2)
        wb.Close(SaveChanges=True)
        excel.Quit()
        log.info("Excel refresh completed")
    except Exception as e:
        log.warning(f"Could not trigger Excel refresh ({e}). Reading existing file.")


def read_loads() -> list[dict]:
    wb = openpyxl.load_workbook(LOADS_FILE, data_only=True, read_only=True)
    ws = wb["Main"]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    now = datetime.now(timezone.utc).isoformat()
    records = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        r = dict(zip(headers, row))
        if not r.get("CE_ID"):
            continue

        records.append({
            "ce_id":               int(r["CE_ID"]),
            "delivery_date":       to_ts(r.get("Delivery_Date")),
            "customer":            clean_str(r.get("Customer")),
            "order_number":        clean_str(r.get("Order_Number")),
            "site_id":             int(r["Site_ID"]) if r.get("Site_ID") is not None else None,
            "site_name":           clean_str(r.get("Site_Name")),
            "terminal_id":         clean_str(r.get("Terminal_ID")),
            "terminal_name":       clean_str(r.get("Terminal_Name")),
            "product_name":        clean_str(r.get("Product_Name")) or "Unknown",
            "gallons_ordered":     float(r["Gallons_Ordered"]) if r.get("Gallons_Ordered") is not None else None,
            "site_address":        clean_str(r.get("Site_Address")),
            "city":                clean_str(r.get("City")),
            "state":               clean_str(r.get("State")),
            "first_name":          clean_str(r.get("First_Name")),
            "last_name":           clean_str(r.get("Last_Name")),
            "start_window":        to_ts(r.get("Start_Window")),
            "end_window":          to_ts(r.get("End_Window")),
            "delivery_eta":        to_ts(r.get("Delivery_ETA")),
            "arrived_at_rack_time":to_ts(r.get("Arrived_at_Rack_Time")),
            "load_status":         int(r["Load_Status"]) if r.get("Load_Status") is not None else 1,
            "customer_id":         int(r["Customer_ID"]) if r.get("Customer_ID") is not None else None,
            "synced_at":           now,
        })

    wb.close()
    return records


def sync(sb: Client):
    records = read_loads()
    if not records:
        log.warning("No load records found in spreadsheet")
        return

    # Deduplicate by (ce_id, product_name) — keep last occurrence
    seen = {}
    for r in records:
        seen[(r["ce_id"], r["product_name"])] = r
    records = list(seen.values())
    log.info(f"Rows after dedup: {len(records)}")

    # Upsert in batches to stay under Supabase request size limits
    batch_size = 500
    total = 0
    for i in range(0, len(records), batch_size):
        batch = records[i:i+batch_size]
        sb.table("loads").upsert(
            batch,
            on_conflict="ce_id,product_name"
        ).execute()
        total += len(batch)

    log.info(f"Sync complete: {total} rows upserted")


def main():
    log.info("Starting sync...")
    trigger_excel_refresh()

    try:
        sb = get_client()
        sync(sb)
    except Exception as e:
        log.error(f"Sync failed: {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
